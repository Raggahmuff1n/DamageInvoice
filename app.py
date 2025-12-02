import streamlit as st
import pandas as pd
from datetime import datetime
import os
from pathlib import Path
import base64
import io
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Page Configuration ---
st.set_page_config(
    page_title="Damage Invoice Tracker", 
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={
        'About': "Damage tracker for legal proceedings"
    }
)

# --- Custom CSS for mobile responsiveness ---
st.markdown("""
<style>
    /* Mobile responsive adjustments */
    @media (max-width: 768px) {
        .block-container {
            padding-left: 1rem;
            padding-right: 1rem;
        }
        .stButton > button {
            width: 100%;
            margin-top: 0.5rem;
        }
    }
    
    /* Make file uploader more prominent */
    .uploadedFile {
        background-color: #f0f8ff;
        border-radius: 5px;
        padding: 10px;
    }
    
    /* Style for clickable links */
    .receipt-link {
        color: #0066cc;
        text-decoration: none;
        font-weight: bold;
    }
    
    .receipt-link:hover {
        text-decoration: underline;
    }
    
    /* Export section styling */
    .export-section {
        background-color: #e8f4f8;
        padding: 1.5rem;
        border-radius: 10px;
        margin-top: 1rem;
        border: 2px solid #1f77b4;
    }
    
    /* Summary cards */
    .summary-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #1f77b4;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# --- Config ---
CATEGORY_LIST = [
    "Property Damage",
    "Economic/Financial Loss", 
    "Medical & Health-Related",
    "Emotional & Psychological Damages",
    "Loss of Companionship or Consortium",
    "Punitive Damages",
    "Special Circumstances",
    "Legal & Administrative Costs",
    "Future Damages",
    "Miscellaneous",
    "Other"
]

# Subcategories for each main category
SUBCATEGORIES = {
    "Property Damage": [
        "Vehicle repair/replacement",
        "Rental vehicle costs",
        "Damage to home or real estate",
        "Damage to personal belongings",
        "Other"
    ],
    "Economic/Financial Loss": [
        "Lost wages or income",
        "Loss of earning capacity",
        "Business interruption",
        "Out-of-pocket expenses",
        "Replacement costs",
        "Other"
    ],
    "Medical & Health-Related": [
        "Medical bills",
        "Medication costs",
        "Rehabilitation or physical therapy",
        "Mental health therapy",
        "Other"
    ],
    "Emotional & Psychological Damages": [
        "Pain and suffering",
        "Emotional distress",
        "Loss of enjoyment of life",
        "Grief and bereavement",
        "Other"
    ],
    "Special Circumstances": [
        "Pet loss and related costs",
        "Temporary housing costs",
        "Childcare expenses",
        "Travel expenses",
        "Other"
    ],
    "Legal & Administrative Costs": [
        "Attorney fees",
        "Court filing fees",
        "Expert witness fees",
        "Other"
    ],
    "Future Damages": [
        "Projected medical care",
        "Future therapy",
        "Long-term disability costs",
        "Other"
    ]
}

# --- Initialize Session State ---
if "damages" not in st.session_state:
    st.session_state["damages"] = []

if "drive_folder_url" not in st.session_state:
    st.session_state["drive_folder_url"] = ""

if "drive_folder_configured" not in st.session_state:
    st.session_state["drive_folder_configured"] = False

if "uploaded_files_data" not in st.session_state:
    st.session_state["uploaded_files_data"] = {}

# --- Helper Functions ---
def save_uploaded_file(uploaded_file):
    """Save uploaded file to session state and return base64 data"""
    bytes_data = uploaded_file.getvalue()
    filename = uploaded_file.name
    
    # Generate unique filename
    timestamp = int(datetime.now().timestamp())
    unique_filename = f"{timestamp}_{filename}"
    
    # Store in session state
    st.session_state["uploaded_files_data"][unique_filename] = {
        'data': bytes_data,
        'original_name': filename,
        'size': len(bytes_data)
    }
    
    return unique_filename

def generate_drive_link(folder_url, filename):
    """Generate expected Google Drive link"""
    if folder_url:
        # Clean the folder URL
        folder_url = folder_url.rstrip('/')
        return f"{folder_url}/{filename}"
    return filename

def create_comprehensive_excel_report(damages_df):
    """Create a comprehensive, lawyer-friendly Excel report with professional formatting"""
    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # Define cell styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        total_font = Font(bold=True, size=12, color="000000")
        total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== SHEET 1: EXECUTIVE SUMMARY =====
        summary_data = []
        summary_data.append(['DAMAGE CLAIM SUMMARY REPORT', '', ''])
        summary_data.append(['', '', ''])
        summary_data.append(['Report Generated:', datetime.now().strftime("%Y-%m-%d %H:%M"), ''])
        summary_data. append(['', '', ''])
        summary_data.append(['KEY METRICS', '', ''])
        summary_data.append(['Total Damages Claimed:', f"${damages_df['Cost'].sum():,.2f}", ''])
        summary_data.append(['Number of Damage Items:', len(damages_df), ''])
        summary_data.append(['Number of Categories:', damages_df['Category'].nunique(), ''])
        summary_data. append(['Average Damage Amount:', f"${damages_df['Cost'].mean():,.2f}", ''])
        summary_data. append(['Highest Single Damage:', f"${damages_df['Cost'].max():,. 2f}", ''])
        summary_data.append(['Lowest Single Damage:', f"${damages_df['Cost'].min():,.2f}", ''])
        summary_data.append(['Date Range:', f"{damages_df['Date'].min()} to {damages_df['Date'].max()}", ''])
        summary_data.append(['', '', ''])
        summary_data.append(['CATEGORY BREAKDOWN', 'Amount', 'Percentage of Total'])
        
        # Calculate category totals and percentages
        total_cost = damages_df['Cost'].sum()
        for category in sorted(damages_df['Category'].unique()):
            cat_total = damages_df[damages_df['Category'] == category]['Cost']. sum()
            percentage = (cat_total / total_cost * 100) if total_cost > 0 else 0
            summary_data.append([category, f"${cat_total:,. 2f}", f"{percentage:.1f}%"])
        
        summary_data.append(['', '', ''])
        summary_data. append(['GRAND TOTAL:', f"${total_cost:,.2f}", '100.0%'])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)
        
        # Format Executive Summary sheet
        ws = writer.sheets['Executive Summary']
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C']. width = 20
        
        # ===== SHEET 2: ALL DAMAGES (CATEGORIZED) =====
        all_damages_data = []
        
        # Sort damages by category, then by date
        sorted_df = damages_df.sort_values(['Category', 'Date'])
        
        current_category = None
        category_totals = {}
        
        for _, row in sorted_df.iterrows():
            if current_category != row['Category']:
                # Add category header
                if current_category is not None:
                    # Add previous category subtotal
                    all_damages_data.append([
                        '', '', '', f'Subtotal for {current_category}:',
                        f"${category_totals[current_category]:,.2f}", '', ''
                    ])
                    all_damages_data.append([''] * 7)  # Empty row
                
                current_category = row['Category']
                category_totals[current_category] = 0
                
                all_damages_data.append([
                    f"CATEGORY: {current_category}", '', '', '', '', '', ''
                ])
                all_damages_data.append([
                    'Date', 'Title', 'Description', 'Amount', 'Receipt File', 'Receipt Link', 'Notes'
                ])
            
            # Add damage entry
            all_damages_data.append([
                row['Date'],
                row['Title'],
                row['Description'],
                f"${row['Cost']:,.2f}",
                row['Receipt'],
                row['Link'],
                ''  # Notes column for lawyer
            ])
            
            category_totals[current_category] += row['Cost']
        
        # Add final category subtotal
        if current_category:
            all_damages_data. append([
                '', '', '', f'Subtotal for {current_category}:',
                f"${category_totals[current_category]:,.2f}", '', ''
            ])
        
        # Add grand total
        all_damages_data.append([''] * 7)
        all_damages_data.append([
            '', '', '', 'GRAND TOTAL:',
            f"${damages_df['Cost'].sum():,. 2f}", '', ''
        ])
        
        all_damages_df = pd.DataFrame(all_damages_data)
        all_damages_df.to_excel(writer, sheet_name='All Damages (Categorized)', index=False, header=False)
        
        # ===== SHEET 3: CATEGORY ANALYSIS =====
        category_analysis = []
        
        for category in sorted(damages_df['Category'].unique()):
            cat_data = damages_df[damages_df['Category'] == category]
            
            category_analysis.append([category, '', '', '', ''])
            category_analysis.append(['Metric', 'Value', '', '', ''])
            category_analysis.append(['Number of Items:', len(cat_data), '', '', ''])
            category_analysis. append(['Total Amount:', f"${cat_data['Cost'].sum():,.2f}", '', '', ''])
            category_analysis.append(['Average Amount:', f"${cat_data['Cost'].mean():,.2f}", '', '', ''])
            category_analysis.append(['Minimum Amount:', f"${cat_data['Cost'].min():,.2f}", '', '', ''])
            category_analysis.append(['Maximum Amount:', f"${cat_data['Cost'].max():,.2f}", '', '', ''])
            category_analysis.append(['Date Range:', f"{cat_data['Date'].min()} to {cat_data['Date'].max()}", '', '', ''])
            category_analysis.append(['% of Total Damages:', f"{(cat_data['Cost'].sum() / total_cost * 100):.1f}%", '', '', ''])
            category_analysis.append(['', '', '', '', ''])  # Empty row between categories
        
        category_analysis_df = pd.DataFrame(category_analysis)
        category_analysis_df.to_excel(writer, sheet_name='Category Analysis', index=False, header=False)
        
        # ===== SHEET 4: MONTHLY BREAKDOWN =====
        if len(damages_df) > 0:
            damages_df['Date_parsed'] = pd.to_datetime(damages_df['Date'])
            damages_df['Year-Month'] = damages_df['Date_parsed'].dt.strftime('%Y-%m')
            
            monthly_data = []
            monthly_data.append(['Month', 'Number of Items', 'Total Amount', 'Average Amount', 'Categories Affected'])
            
            for month in sorted(damages_df['Year-Month'].unique()):
                month_data = damages_df[damages_df['Year-Month'] == month]
                categories = ', '.join(month_data['Category'].unique())
                
                monthly_data.append([
                    month,
                    len(month_data),
                    f"${month_data['Cost'].sum():,.2f}",
                    f"${month_data['Cost'].mean():,.2f}",
                    categories
                ])
            
            monthly_data. append(['', '', '', '', ''])
            monthly_data. append([
                'TOTAL',
                len(damages_df),
                f"${damages_df['Cost'].sum():,.2f}",
                f"${damages_df['Cost'].mean():,. 2f}",
                f"{damages_df['Category'].nunique()} categories"
            ])
            
            monthly_df = pd.DataFrame(monthly_data)
            monthly_df.to_excel(writer, sheet_name='Monthly Breakdown', index=False, header=False)
        
        # ===== SHEET 5: RECEIPT CHECKLIST =====
        receipt_data = []
        receipt_data. append(['Receipt Checklist for Legal Documentation', '', '', ''])
        receipt_data. append(['', '', '', ''])
        receipt_data.append(['☐', 'Date', 'Item', 'Receipt Status'])
        
        for _, row in sorted_df.iterrows():
            receipt_status = 'Uploaded' if row['Receipt'] else 'Missing'
            receipt_data.append(['☐', row['Date'], row['Title'], receipt_status])
        
        receipt_data.append(['', '', '', ''])
        receipt_data.append(['Total Items:', len(damages_df), '', ''])
        receipt_data.append(['With Receipts:', len(damages_df[damages_df['Receipt'] != '']), '', ''])
        receipt_data.append(['Missing Receipts:', len(damages_df[damages_df['Receipt'] == '']), '', ''])
        
        receipt_df = pd. DataFrame(receipt_data)
        receipt_df.to_excel(writer, sheet_name='Receipt Checklist', index=False, header=False)
        
        # ===== SHEET 6: SUPPORTING DOCUMENTS =====
        docs_data = []
        docs_data.append(['Supporting Documents Location', ''])
        docs_data.append(['', ''])
        docs_data.append(['Google Drive Folder:', st.session_state. get("drive_folder_url", "Not configured")])
        docs_data.append(['', ''])
        docs_data.append(['Instructions for Access:', ''])
        docs_data.append(['1. Click on the Google Drive folder link above', ''])
        docs_data. append(['2. All receipt images are stored in this folder', ''])
        docs_data.append(['3. File names match the "Receipt File" column in the damage list', ''])
        docs_data.append(['4.  Ensure you have view access to the folder', ''])
        
        docs_df = pd.DataFrame(docs_data)
        docs_df.to_excel(writer, sheet_name='Supporting Documents', index=False, header=False)
    
    return output.getvalue()

def create_legal_summary_document(damages_df):
    """Create a formatted text document for legal proceedings"""
    total_cost = damages_df['Cost'].sum()
    
    summary = f"""
================================================================================
                        DAMAGE CLAIM DOCUMENTATION
                           LEGAL SUMMARY REPORT
================================================================================

REPORT GENERATED: {datetime.now().strftime('%Y-%m-%d at %H:%M')}

--------------------------------------------------------------------------------
I. EXECUTIVE SUMMARY
--------------------------------------------------------------------------------

TOTAL DAMAGES CLAIMED: ${total_cost:,.2f}

This document provides a comprehensive summary of all damages incurred, 
organized by category for legal proceedings.

Key Statistics:
• Total Number of Damage Items: {len(damages_df)}
• Number of Categories: {damages_df['Category'].nunique()}
• Date Range: {damages_df['Date'].min()} to {damages_df['Date'].max()}
• Average Damage Amount: ${damages_df['Cost']. mean():,.2f}
• Highest Single Damage: ${damages_df['Cost'].max():,. 2f}
• Lowest Single Damage: ${damages_df['Cost'].min():,.2f}

--------------------------------------------------------------------------------
II. DAMAGE BREAKDOWN BY CATEGORY
--------------------------------------------------------------------------------
"""
    
    # Add detailed breakdown by category
    for category in sorted(damages_df['Category'].unique()):
        cat_data = damages_df[damages_df['Category'] == category]
        cat_total = cat_data['Cost'].sum()
        percentage = (cat_total / total_cost * 100) if total_cost > 0 else 0
        
        summary += f"""
{category. upper()}
{'=' * len(category)}
Total: ${cat_total:,.2f} ({percentage:.1f}% of total damages)
Number of Items: {len(cat_data)}
Average per Item: ${cat_data['Cost'].mean():,.2f}

Items:
"""
        for _, row in cat_data.iterrows():
            summary += f"  • {row['Date']} - {row['Title']}: ${row['Cost']:,.2f}\n"
            if row['Description']:
                summary += f"    Description: {row['Description']}\n"
            if row['Receipt']:
                summary += f"    Receipt: {row['Receipt']}\n"
    
    summary += f"""
--------------------------------------------------------------------------------
III.  CHRONOLOGICAL LISTING
--------------------------------------------------------------------------------

"""
    # Add chronological listing
    sorted_by_date = damages_df.sort_values('Date')
    for _, row in sorted_by_date.iterrows():
        summary += f"{row['Date']} | {row['Category']} | {row['Title']} | ${row['Cost']:,.2f}\n"
    
    summary += f"""
--------------------------------------------------------------------------------
IV. RECEIPT DOCUMENTATION STATUS
--------------------------------------------------------------------------------

Total Items: {len(damages_df)}
Items with Receipts: {len(damages_df[damages_df['Receipt'] != ''])}
Items Missing Receipts: {len(damages_df[damages_df['Receipt'] == ''])}

Receipt Storage Location:
{st.session_state.get('drive_folder_url', 'Not configured')}

--------------------------------------------------------------------------------
V. TOTAL DAMAGES SUMMARY
--------------------------------------------------------------------------------

GRAND TOTAL OF ALL DAMAGES: ${total_cost:,.2f}

This amount represents the total of all documented damages with supporting 
evidence as detailed above.

--------------------------------------------------------------------------------
                           END OF REPORT
--------------------------------------------------------------------------------
"""
    
    return summary

# --- Main App ---
st.title("⚖️ Damage Invoice Tracker")
st.markdown("### Legal Proceedings Documentation System")

# --- Configuration Section ---
with st.expander("⚙️ Configure Google Drive Folder (One-time Setup)", expanded=not st.session_state["drive_folder_configured"]):
    st.markdown("""
    📁 **Set your Google Drive folder where receipts will be stored**
    
    1. Create or open a folder in Google Drive
    2. Right-click → "Get link" → Set to "Anyone with the link can view"
    3. Copy the folder URL and paste below
    """)
    
    drive_url = st.text_input(
        "Google Drive Folder URL:",
        value=st.session_state.get("drive_folder_url", "https://drive.google.com/drive/folders/1HIu5XR7pFg8s49AG8Yiu_7aCK9HMeT8N"),
        help="Example: https://drive.google.com/drive/folders/..."
    )
    
    col1, col2 = st. columns([1, 3])
    with col1:
        if st.button("Save Configuration", type="primary", use_container_width=True):
            st.session_state["drive_folder_url"] = drive_url
            st.session_state["drive_folder_configured"] = True
            st.success("✅ Configuration saved!")
            st.rerun()
    
    with col2:
        st.info("ℹ️ You'll manually upload files to this folder. The app generates the links.")

if st.session_state["drive_folder_configured"]:
    st. success(f"✅ Connected to folder: `{st.session_state['drive_folder_url']}`")

# --- Entry Form ---
st.markdown("---")
st. subheader("📝 Add New Damage Entry")

with st. form("damage_form", clear_on_submit=True):
    # Responsive columns
    col1, col2 = st.columns([1, 1])
    
    with col1:
        title = st.text_input(
            "Title *",
            placeholder="Brief description",
            key="form_title"
        )
        
        category = st.selectbox(
            "Category *",
            CATEGORY_LIST,
            key="form_category"
        )
        
        # Show subcategories if available
        subcategory = ""
        custom_subcategory = ""
        if category in SUBCATEGORIES:
            subcategory = st.selectbox(
                f"Subcategory",
                ["Select... "] + SUBCATEGORIES[category],
                key="form_subcategory"
            )
            
            # Custom input for "Other" subcategory
            if subcategory == "Other":
                custom_subcategory = st.text_input(
                    "Specify:",
                    placeholder="Enter custom subcategory",
                    key="form_custom_subcategory"
                )
        
        # Custom input for "Other" main category
        custom_category = ""
        if category == "Other":
            custom_category = st.text_input(
                "Specify category:",
                placeholder="Enter custom category",
                key="form_custom_category"
            )
    
    with col2:
        date = st.date_input(
            "Date *",
            value=datetime.today(),
            key="form_date"
        )
        
        cost = st.number_input(
            "Cost (USD) *",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            value=0.0,
            key="form_cost"
        )
        
        description = st. text_area(
            "Description",
            height=70,
            placeholder="Additional details (optional)",
            key="form_description"
        )
    
    # File upload
    image_file = st.file_uploader(
        "📎 Upload Receipt/Invoice",
        type=["png", "jpg", "jpeg", "pdf"],
        help="Upload supporting documentation",
        key="form_file"
    )
    
    # Submit button
    submitted = st.form_submit_button(
        "➕ Add Damage Entry",
        type="primary",
        use_container_width=True
    )

# --- Handle Form Submission ---
if submitted:
    if not title:
        st.error("❌ Please provide a title")
    elif cost <= 0:
        st. error("❌ Please enter a valid cost amount")
    else:
        # Handle file upload
        filename = ""
        file_link = ""
        
        if image_file is not None:
            # Save file to session state
            filename = save_uploaded_file(image_file)
            
            # Generate Google Drive link
            if st.session_state["drive_folder_configured"]:
                file_link = generate_drive_link(st.session_state["drive_folder_url"], filename)
            else:
                file_link = filename
        
        # Determine final category
        if category == "Other":
            final_category = custom_category if custom_category else "Other"
        elif category in SUBCATEGORIES and subcategory not in ["Select...", None, ""]:
            if subcategory == "Other":
                subcategory_text = custom_subcategory if custom_subcategory else "Other"
            else:
                subcategory_text = subcategory
            final_category = f"{category} - {subcategory_text}"
        else:
            final_category = category
        
        # Create entry
        entry = {
            "Title": title,
            "Description": description,
            "Date": date.strftime("%Y-%m-%d"),
            "Category": final_category,
            "Cost": float(cost),
            "Receipt": filename,
            "Link": file_link
        }
        
        st.session_state["damages"].append(entry)
        
        # Show success message
        st.success("✅ Damage entry added successfully!")
        
        if image_file and st.session_state["drive_folder_configured"]:
            st.info(f"""
            📤 **Next Step:** Upload `{filename}` to your Google Drive folder:
            1. Download the file below
            2. Upload to your Google Drive folder
            3. The link in the spreadsheet will work automatically
            """)
            
            # Provide download link
            if filename in st.session_state["uploaded_files_data"]:
                file_data = st.session_state["uploaded_files_data"][filename]['data']
                st.download_button(
                    label=f"⬇️ Download {filename} for Google Drive upload",
                    data=file_data,
                    file_name=filename,
                    mime="application/octet-stream",
                    key=f"download_{filename}"
                )
        
        # Trigger rerun to clear form
        st.rerun()

# --- Display and Export Section ---
st.markdown("---")
st.header("📊 Damage Summary & Legal Documentation Export")

if st.session_state["damages"]:
    df = pd.DataFrame(st.session_state["damages"])
    
    # Calculate key metrics
    total_cost = df["Cost"].sum()
    
    # Display comprehensive summary
    st.markdown("### 💼 Case Summary")
    
    # Main metrics in columns
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("💰 **TOTAL DAMAGES**", f"${total_cost:,.2f}", 
                 help="Total amount of all damages claimed")
    with col2:
        st.metric("📝 Total Items", len(df),
                 help="Number of individual damage entries")
    with col3:
        st.metric("📊 Average Damage", f"${df['Cost'].mean():,.2f}",
                 help="Average cost per damage item")
    with col4:
        st.metric("📁 Categories", df["Category"].nunique(),
                 help="Number of different damage categories")
    
    # Category breakdown with visual cards
    st.markdown("### 📂 Damage Categories Breakdown")
    
    # Create category summary with percentages
    category_summary = []
    for category in sorted(df['Category'].unique()):
        cat_data = df[df['Category'] == category]
        cat_total = cat_data['Cost']. sum()
        percentage = (cat_total / total_cost * 100) if total_cost > 0 else 0
        category_summary.append({
            'Category': category,
            'Total': cat_total,
            'Count': len(cat_data),
            'Percentage': percentage,
            'Average': cat_data['Cost'].mean()
        })
    
    # Display categories in expandable sections
    for cat_info in category_summary:
        with st.expander(f"**{cat_info['Category']}** - ${cat_info['Total']:,.2f} ({cat_info['Percentage']:.1f}%)"):
            col1, col2, col3 = st.columns(3)
            with col1:
                st.write(f"**Total:** ${cat_info['Total']:,.2f}")
            with col2:
                st.write(f"**Items:** {cat_info['Count']}")
            with col3:
                st.write(f"**Average:** ${cat_info['Average']:,.2f}")
            
            # Show items in this category
            cat_items = df[df['Category'] == cat_info['Category']][['Date', 'Title', 'Cost', 'Receipt']]
            for _, item in cat_items.iterrows():
                receipt_status = "✅" if item['Receipt'] else "❌"
                st.write(f"• {item['Date']} - {item['Title']}: **${item['Cost']:,.2f}** {receipt_status}")
    
    # Export Section - Prominently displayed
    st.markdown("---")
    st.markdown("### 📥 **Export for Attorney / Legal Proceedings**")
    st.markdown('<div class="export-section">', unsafe_allow_html=True)
    
    st.markdown("#### 📋 Download Complete Legal Documentation Package:")
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col1:
        # Comprehensive Excel Report
        excel_data = create_comprehensive_excel_report(df)
        st.download_button(
            "📊 **Download Complete Excel Report**",
            data=excel_data,
            file_name=f"Legal_Damage_Report_{datetime. now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Comprehensive Excel report with 6 sheets including executive summary, categorized damages, analysis, and receipt checklist"
        )
        st.caption("✓ Executive Summary\n✓ Categorized Damages\n✓ Category Analysis\n✓ Monthly Breakdown\n✓ Receipt Checklist")
    
    with col2:
        # Legal Summary Document
        legal_summary = create_legal_summary_document(df)
        st.download_button(
            "📄 **Download Legal Summary**",
            data=legal_summary,
            file_name=f"Legal_Summary_{datetime.now().strftime('%Y%m%d')}. txt",
            mime="text/plain",
            use_container_width=True,
            help="Formatted text document suitable for legal proceedings"
        )
        st.caption("✓ Executive Summary\n✓ Category Breakdown\n✓ Chronological List\n✓ Receipt Status\n✓ Total Damages")
    
    with col3:
        # Simple CSV for database import
        csv_data = df. to_csv(index=False). encode('utf-8')
        st.download_button(
            "📈 **Download CSV Data**",
            data=csv_data,
            file_name=f"Damage_Data_{datetime.now(). strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
            help="Simple CSV format for database import or additional analysis"
        )
        st.caption("✓ Raw Data Format\n✓ Import to Any System\n✓ Maximum Compatibility")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Attorney instructions
    st.info("""
    📨 **Instructions for Your Attorney:**
    
    1.  **Download the Complete Excel Report** - This contains all damage information organized in 6 comprehensive sheets
    2. **Review the Executive Summary** sheet for a quick overview of total damages
    3. **Check the "All Damages (Categorized)"** sheet for detailed itemization with subtotals
    4. **Access receipts** via the Google Drive folder link in the "Supporting Documents" sheet
    5. **Use the Legal Summary** document for quick reference during proceedings
    
    **Google Drive Receipts Folder:** {drive_link}
    """.format(drive_link=st.session_state.get('drive_folder_url', 'Not configured')))
    
    # Additional Analysis Tabs
    st.markdown("---")
    st.markdown("### 📊 Detailed Analysis")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📋 All Entries", "📊 Visual Analysis", "📅 Timeline", "📎 Receipt Status"])
    
    with tab1:
        # Sortable, searchable dataframe
        st.subheader("Complete Damage List")
        display_df = df.copy()
        display_df['Cost'] = display_df['Cost']. apply(lambda x: f"${x:,.2f}")
        display_df['Has Receipt'] = display_df['Receipt'].apply(lambda x: '✅' if x else '❌')
        st.dataframe(
            display_df[['Date', 'Category', 'Title', 'Description', 'Cost', 'Has Receipt']],
            use_container_width=True,
            height=400
        )
    
    with tab2:
        # Visual breakdown
        st.subheader("Visual Category Breakdown")
        
        # Pie chart of categories
        category_totals = df.groupby('Category')['Cost'].sum().sort_values(ascending=False)
        
        # Bar chart
        st.bar_chart(category_totals)
        
        # Summary statistics
        st.subheader("Statistical Summary")
        col1, col2 = st.columns(2)
        with col1:
            st.write("**Top 5 Highest Damages:**")
            top_5 = df.nlargest(5, 'Cost')[['Title', 'Cost']]
            for _, item in top_5.iterrows():
                st.write(f"• {item['Title']}: ${item['Cost']:,.2f}")
        
        with col2:
            st.write("**Category Rankings:**")
            for i, (cat, total) in enumerate(category_totals.head(5).items(), 1):
                st.write(f"{i}. {cat}: ${total:,.2f}")
    
    with tab3:
        # Timeline view
        st.subheader("Damage Timeline")
        
        # Group by month
        df['Date_parsed'] = pd.to_datetime(df['Date'])
        df['Month'] = df['Date_parsed']. dt.strftime('%Y-%m')
        
        monthly_totals = df.groupby('Month')['Cost'].sum()
        st.line_chart(monthly_totals)
        
        # Monthly breakdown table
        monthly_summary = df.groupby('Month'). agg({
            'Cost': ['count', 'sum', 'mean']
        }).round(2)
        monthly_summary.columns = ['Count', 'Total ($)', 'Average ($)']
        st.dataframe(monthly_summary, use_container_width=True)
    
    with tab4:
        # Receipt tracking
        st.subheader("Receipt Documentation Status")
        
        receipts_uploaded = len(df[df['Receipt'] != ''])
        receipts_missing = len(df[df['Receipt'] == ''])
        
        col1, col2 = st.columns(2)
        with col1:
            st. metric("✅ Receipts Uploaded", receipts_uploaded)
        with col2:
            st.metric("❌ Receipts Missing", receipts_missing)
        
        # List items missing receipts
        if receipts_missing > 0:
            st.warning(f"**{receipts_missing} items are missing receipts:**")
            missing_receipts = df[df['Receipt'] == ''][['Date', 'Title', 'Cost']]
            for _, item in missing_receipts.iterrows():
                st.write(f"• {item['Date']} - {item['Title']}: ${item['Cost']:,.2f}")
        else:
            st.success("✅ All items have receipts uploaded!")
        
        # Google Drive folder reminder
        if st.session_state["drive_folder_configured"]:
            st.info(f"📁 **Receipt Storage:** [{st.session_state['drive_folder_url']}]({st.session_state['drive_folder_url']})")

else:
    st.info("No damages recorded yet. Add your first entry using the form above.")

# --- Instructions ---
with st.expander("📖 How to Use for Legal Documentation"):
    st.markdown("""
    ### For Legal Proceedings:
    
    **Step 1: Document All Damages**
    - Add each damage with detailed information
    - Upload all receipts and invoices
    - Be specific in descriptions
    
    **Step 2: Review and Verify**
    - Check the category breakdown for accuracy
    - Ensure all receipts are uploaded
    - Verify dates and amounts
    
    **Step 3: Export for Attorney**
    - Download the **Complete Excel Report** (primary document)
    - Download the **Legal Summary** for quick reference
    - Share your Google Drive folder with view access
    
    **What Your Attorney Receives:**
    1. **Executive Summary** with total damages and breakdown
    2. **Itemized list** organized by category with subtotals
    3. **Category analysis** showing impact of each damage type
    4. **Monthly timeline** of damages incurred
    5. **Receipt checklist** for verification
    6. **Direct links** to all supporting documents
    
    **Tips for Legal Use:**
    - Keep descriptions factual and concise
    - Upload clear, legible receipts
    - Maintain chronological accuracy
    - Document everything, no matter how small
    """)

# --- Footer ---
st.markdown("---")
st.caption("Damage Invoice Tracker v3.0 | Professional Legal Documentation System")
